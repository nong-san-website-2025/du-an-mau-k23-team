from rest_framework import viewsets, permissions, generics, status
from rest_framework.views import APIView 
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, IsAdminUser, AllowAny
from django.db.models import Q, F
from django.utils.timezone import now
from django.utils import timezone
from django.db import transaction
from django.contrib.auth import get_user_model
from rest_framework import filters
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .excel_import_service import FlashSaleExcelImportService
from datetime import datetime, timedelta
import traceback 

from sellers.models import Seller
from .models import Voucher, FlashSale, UserVoucher

from .serializers import (
    VoucherDetailSerializer,
    FlashSaleSerializer,
    FlashSaleAdminSerializer,
    UserVoucherSerializer,
    SellerVoucherSerializer,
    VoucherImportSerializer 
)

User = get_user_model()

# ============================================================
# 1. HELPER CLASSES & FUNCTIONS
# ============================================================

class IsStaffOrReadOnly(permissions.BasePermission):
    def has_permission(self, request, view):
        
        if request.method in permissions.SAFE_METHODS:
            return request.user and request.user.is_authenticated
        return request.user and request.user.is_staff

def _safe_get_user_voucher_by_code(user, code, for_update=False):
    qs = UserVoucher.objects.select_related("voucher").filter(user=user, voucher__code=code).order_by("id")
    if for_update:
        qs = qs.select_for_update()
    return qs.first()

def _calc_discount_for_voucher(voucher, order_total):
    discount = 0.0
    if voucher.discount_amount:
        discount = float(voucher.discount_amount)
    elif voucher.discount_percent:
        discount = (float(order_total) * float(voucher.discount_percent)) / 100.0
        if voucher.max_discount_amount and discount > float(voucher.max_discount_amount):
            discount = float(voucher.max_discount_amount)
    elif voucher.freeship_amount:
        discount = float(voucher.freeship_amount)
    return float(discount)


# ============================================================
# 2. VOUCHER VIEWSET
# ============================================================

class VoucherViewSet(viewsets.ModelViewSet):
    queryset = Voucher.objects.all().order_by('-created_at')
    serializer_class = VoucherDetailSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if not user.is_staff:
            seller = getattr(user, 'seller', None) or getattr(user, 'store', None)
            if seller:
                return qs.filter(Q(scope="system") | Q(seller=seller))
            return qs.filter(scope="system")
        return qs

    def perform_create(self, serializer):
        user = self.request.user
        if serializer.validated_data.get('scope') == 'system' and not user.is_staff:
            raise PermissionError("Chỉ admin mới tạo voucher hệ thống.")
        voucher = serializer.save(created_by=user)

        if voucher.distribution_type == Voucher.DistributionType.DIRECT:
            users_qs = User.objects.filter(is_active=True)
            remaining = voucher.total_quantity
            bulk = []
            for u in users_qs:
                if remaining is not None and remaining <= 0: break
                give = voucher.per_user_quantity
                if remaining is not None: give = min(give, remaining)
                if not UserVoucher.objects.filter(user=u, voucher=voucher).exists():
                    bulk.append(UserVoucher(user=u, voucher=voucher, quantity=give))
                    if remaining is not None: remaining -= give
            if bulk:
                UserVoucher.objects.bulk_create(bulk, ignore_conflicts=True)


# ============================================================
# 3. IMPORT VOUCHER API (FIX LỖI 500 FREESHIP)
# ============================================================

class ImportVoucherAPIView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request):
        try:
            data = request.data
            if not isinstance(data, list):
                return Response({"error": "Dữ liệu gửi lên phải là danh sách (Array)."}, status=400)

            serializer = VoucherImportSerializer(data=data, many=True)
            if not serializer.is_valid():
                return Response({"errors": serializer.errors, "message": "Dữ liệu không hợp lệ."}, status=400)

            validated_data = serializer.validated_data
            vouchers_to_create = []
            current_user = request.user if request.user.is_authenticated else None

            with transaction.atomic():
                for item in validated_data:
                    start_at = datetime.datetime.combine(item['start_date'], datetime.time.min)
                    end_at = datetime.datetime.combine(item['end_date'], datetime.time.max)
                    if timezone.is_naive(start_at): start_at = timezone.make_aware(start_at)
                    if timezone.is_naive(end_at): end_at = timezone.make_aware(end_at)

                    # --- Logic Phân loại giá trị ---
                    d_percent = 0
                    d_amount = 0
                    d_freeship = 0
                    max_disc = None # Để None thay vì 0 nếu DB cho phép null

                    # Chuẩn hóa input
                    raw_type = str(item['discount_type']).lower().strip()

                    if raw_type == 'percent':
                        d_percent = item['value']
                        max_disc = 500000 
                    elif raw_type == 'freeship':
                        d_freeship = item['value']
                    else: 
                        d_amount = item['value']

                    # --- [FIX QUAN TRỌNG] Bỏ trường voucher_type vì DB không có ---
                    voucher = Voucher(
                        code=item['code'],
                        title=item['title'],
                        
                        # Chỉ lưu các giá trị tiền/%, không lưu voucher_type text
                        discount_percent=d_percent,
                        discount_amount=d_amount,
                        freeship_amount=d_freeship,
                        max_discount_amount=max_disc,
                        
                        min_order_value=item.get('min_order', 0),
                        start_at=start_at,
                        end_at=end_at,
                        
                        total_quantity=item['quantity'],
                        per_user_quantity=1,
                        
                        # Mặc định là Claim để user nhận được
                        distribution_type='claim', 
                        
                        scope='system', 
                        active=True,
                        created_by=current_user
                    )
                    vouchers_to_create.append(voucher)
                
                Voucher.objects.bulk_create(vouchers_to_create)

            return Response({"success": True, "message": "Import thành công!"}, status=201)

        except Exception as e:
            # In lỗi ra terminal để debug
            print("❌ LỖI IMPORT 500:", str(e))
            print(traceback.format_exc())
            return Response({"error": f"Lỗi hệ thống: {str(e)}"}, status=500)


# ============================================================
# 4. API CLAIM VOUCHER
# ============================================================

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def claim_voucher(request):
    code = request.data.get("code")
    user = request.user

    if not code:
        return Response({"error": "Thiếu mã voucher"}, status=400)

    try:
        with transaction.atomic():
            voucher = Voucher.objects.select_for_update().filter(
                code=code, active=True
            ).order_by('id').first()

            if not voucher:
                return Response({"error": "Voucher không tồn tại"}, status=404)

            if voucher.start_at and now() < voucher.start_at:
                return Response({"error": "Voucher chưa bắt đầu"}, status=400)
            if voucher.end_at and now() > voucher.end_at:
                return Response({"error": "Voucher đã hết hạn"}, status=400)

            if UserVoucher.objects.filter(user=user, voucher=voucher).exists():
                return Response({"error": "Bạn đã sở hữu voucher này rồi"}, status=400)

            # Check số lượng (Hỗ trợ vô cực)
            total_qty = voucher.total_quantity
            used_qty = voucher.used_quantity if hasattr(voucher, 'used_quantity') else 0
            if used_qty == 0 and hasattr(voucher, 'issued_count'):
                 used_qty = voucher.issued_count()

            if total_qty is not None:
                if used_qty >= total_qty:
                    return Response({"error": "Voucher đã hết lượt sử dụng"}, 400)

            uv = UserVoucher.objects.create(
                user=user, 
                voucher=voucher, 
                quantity=voucher.per_user_quantity or 1
            )
            
            if hasattr(voucher, 'used_quantity'):
                Voucher.objects.filter(pk=voucher.pk).update(used_quantity=F('used_quantity') + 1)

    except Exception as e:
        return Response({"error": str(e)}, status=400)

    serializer = UserVoucherSerializer(uv)
    return Response({"success": True, "message": "Nhận voucher thành công!", "user_voucher": serializer.data})


# ============================================================
# 5. API OVERVIEW (FILTER)
# ============================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def promotions_overview(request):
    user = request.user
    search = request.query_params.get("search")
    scope = request.query_params.get("scope")
    distribution_type = request.query_params.get("distribution_type")
    active = request.query_params.get("active")
    voucher_type_filter = request.query_params.get("voucherType")

    if user.is_staff:
        vouchers = Voucher.objects.all().order_by('-created_at')
    else:
        seller = getattr(user, 'seller', None) or getattr(user, 'store', None)
        if seller:
            vouchers = Voucher.objects.filter(Q(scope='system') | Q(seller=seller)).order_by('-created_at')
        else:
            vouchers = Voucher.objects.filter(scope='system').order_by('-created_at')

    if search:
        vouchers = vouchers.filter(Q(title__icontains=search) | Q(code__icontains=search))
    
    # Logic Lọc chuẩn (dựa vào giá trị tiền, không dựa vào cột text)
    if voucher_type_filter:
        if voucher_type_filter == 'freeship':
            vouchers = vouchers.filter(freeship_amount__gt=0)
        elif voucher_type_filter == 'normal':
            vouchers = vouchers.filter(Q(discount_amount__gt=0) | Q(discount_percent__gt=0))

    if scope: vouchers = vouchers.filter(scope=scope)
    if distribution_type: vouchers = vouchers.filter(distribution_type=distribution_type)
    if active is not None:
        if active.lower() in ["true", "1"]: vouchers = vouchers.filter(active=True)
        elif active.lower() in ["false", "0"]: vouchers = vouchers.filter(active=False)

    data = []
    for v in vouchers:
        d_type = getattr(v, 'discount_type', 'unknown')
        if callable(d_type): d_type = d_type()

        used = 0
        if hasattr(v, 'used_quantity'): used = v.used_quantity
        elif hasattr(v, 'issued_count'): 
            m = getattr(v, 'issued_count')
            used = m() if callable(m) else m

        total = v.total_quantity
        remaining = max(0, total - used) if total is not None else None

        # Determine type for Frontend
        final_type = 'normal'
        if v.freeship_amount and v.freeship_amount > 0:
            final_type = 'freeship'

        data.append({
            "id": v.id,    
            "code": v.code,
            "name": v.title or v.code,
            "type": "voucher",
            "voucher_type": final_type, # Frontend dùng cái này để hiển thị Tag
            "discount_type": d_type, 
            "discount_percent": float(v.discount_percent) if v.discount_percent else None,
            "discount_amount": int(v.discount_amount) if v.discount_amount else None,
            "freeship_amount": int(v.freeship_amount) if v.freeship_amount else None,
            "min_order_value": int(v.min_order_value) if v.min_order_value else 0,
            "start": v.start_at,
            "end": v.end_at,
            "scope": v.scope,
            "active": v.active,
            "distribution_type": v.distribution_type,
            "total_quantity": total,
            "per_user_quantity": v.per_user_quantity,
            "issued_count": used,
            "remaining_quantity": remaining,
        })
    return Response(data)


# ============================================================
# 6. OTHER VIEWS (GIỮ NGUYÊN)
# ============================================================

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def my_vouchers(request):
    user_vouchers = UserVoucher.objects.filter(user=request.user).select_related('voucher')
    serializer = UserVoucherSerializer(user_vouchers, many=True)
    return Response(serializer.data)

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def apply_voucher(request):
    code = request.data.get("code")
    order_total = request.data.get("order_total")
    if code is None or order_total is None:
        return Response({"error": "Thiếu code hoặc order_total"}, status=400)
    user_voucher = _safe_get_user_voucher_by_code(request.user, code, for_update=False)
    if not user_voucher:
        return Response({"error": "Voucher không thuộc về bạn"}, status=400)
    voucher = user_voucher.voucher
    if not voucher.active: return Response({"error": "Voucher đã tắt"}, status=400)
    if voucher.start_at and now() < voucher.start_at: return Response({"error": "Chưa đến thời gian"}, status=400)
    if voucher.end_at and now() > voucher.end_at: return Response({"error": "Hết hạn"}, status=400)
    
    # Check usage user
    remaining_user = 0
    if hasattr(user_voucher, 'remaining_for_user'): remaining_user = user_voucher.remaining_for_user()
    else: remaining_user = max(0, user_voucher.quantity - user_voucher.used_count)
    if remaining_user <= 0: return Response({"error": "Đã dùng hết"}, status=400)

    try: order_total_f = float(order_total)
    except: return Response({"error": "order_total sai"}, status=400)
    
    if voucher.min_order_value and order_total_f < float(voucher.min_order_value):
        return Response({"error": f"Đơn tối thiểu {voucher.min_order_value}₫"}, status=400)

    discount = _calc_discount_for_voucher(voucher, order_total_f)
    new_total = max(0.0, order_total_f - discount)
    return Response({"success": True, "discount": discount, "new_total": new_total})

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def consume_voucher(request):
    code = request.data.get("code")
    order_total = request.data.get("order_total")
    if code is None or order_total is None:
        return Response({"error": "Thiếu code hoặc order_total"}, status=400)
    try:
        with transaction.atomic():
            user_voucher = _safe_get_user_voucher_by_code(request.user, code, for_update=True)
            if not user_voucher: return Response({"error": "Lỗi voucher user"}, status=400)
            voucher = user_voucher.voucher
            if not voucher.active: return Response({"error": "Voucher tắt"}, 400)
            
            # Recalculate check
            remaining_user = max(0, user_voucher.quantity - user_voucher.used_count)
            if remaining_user <= 0: return Response({"error": "Hết lượt"}, 400)

            discount = _calc_discount_for_voucher(voucher, float(order_total))
            if hasattr(user_voucher, 'mark_used_once'): user_voucher.mark_used_once()
            else:
                user_voucher.used_count += 1
                user_voucher.save()
            
            if hasattr(voucher, 'used_quantity'):
                Voucher.objects.filter(pk=voucher.pk).update(used_quantity=F('used_quantity') + 1)
            return Response({"success": True, "discount": discount})
    except Exception as e:
        return Response({"error": str(e)}, status=400)

class FlashSaleListView(generics.ListCreateAPIView):
    serializer_class = FlashSaleSerializer
    permission_classes = [AllowAny]
    def get_queryset(self):
        now_t = timezone.now()
        return FlashSale.objects.filter(is_active=True, start_time__lte=now_t, end_time__gt=now_t).prefetch_related('flashsale_products__product')

class FlashSaleAdminViewSet(viewsets.ModelViewSet):
    queryset = FlashSale.objects.all().prefetch_related('products')
    serializer_class = FlashSaleAdminSerializer
    permission_classes = [IsAdminUser]

class IsSellerUser(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated and hasattr(request.user, 'seller')

class SellerVoucherViewSet(viewsets.ModelViewSet):
    serializer_class = SellerVoucherSerializer
    permission_classes = [IsSellerUser]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['active']
    search_fields = ['title', 'code']
    def get_queryset(self):
        return Voucher.objects.filter(seller=self.request.user.seller).order_by('-created_at')
    def perform_create(self, serializer):
        scope_val = "seller" 
        if hasattr(Voucher, 'Scope'): scope_val = Voucher.Scope.SELLER
        serializer.save(created_by=self.request.user, seller=self.request.user.seller, scope=scope_val)

@api_view(['GET'])
@permission_classes([AllowAny])
def public_seller_vouchers(request, seller_id):
    try: seller = Seller.objects.get(id=seller_id)
    except: return Response({"error": "Seller 404"}, 404)
    now = timezone.now()
    vouchers = Voucher.objects.filter(
        seller=seller,
        active=True,
        scope=Voucher.Scope.SELLER
    ).filter(
        Q(start_at__isnull=True) | Q(start_at__lte=now)
    ).filter(
        Q(end_at__isnull=True) | Q(end_at__gte=now)
    ).order_by('-created_at')

    # Serialize dữ liệu (dùng VoucherDetailSerializer hoặc tạo serializer đơn giản hơn)
    serializer = VoucherDetailSerializer(vouchers, many=True)
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([IsAdminUser])
def import_flash_sale_excel(request):
    """
    API endpoint để import Flash Sale từ file Excel
    """
    if 'file' not in request.FILES:
        return Response({'error': 'File không được tìm thấy'}, status=400)
    
    file = request.FILES['file']
    service = FlashSaleExcelImportService()
    
    if service.import_flash_sales(file):
        result = service.get_result()
        return Response(result, status=201)
    else:
        result = service.get_result()
        return Response(result, status=400)


@api_view(['GET'])
@permission_classes([IsAdminUser])
def download_flash_sale_template(request):
    """
    API endpoint để download template Excel cho Flash Sale import
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Flash Sales"
    
    headers = ['product_id', 'product_name', 'flash_price', 'stock', 'start_time', 'end_time']
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    now_time = datetime.now()
    sample_start = now_time + timedelta(days=1)
    sample_end = sample_start + timedelta(hours=2)
    
    sample_data = [
        [1, 'Sản phẩm mẫu 1', 50000, 100, sample_start.strftime('%Y-%m-%d %H:%M:%S'), sample_end.strftime('%Y-%m-%d %H:%M:%S')],
        [2, 'Sản phẩm mẫu 2', 75000, 50, sample_start.strftime('%Y-%m-%d %H:%M:%S'), sample_end.strftime('%Y-%m-%d %H:%M:%S')],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 22
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=flash_sale_template.xlsx'
    
    wb.save(response)
    return response